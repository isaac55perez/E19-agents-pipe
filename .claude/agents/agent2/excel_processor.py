"""
Excel file reading and writing functionality.

This module provides:
- Reading repository data from Excel files
- Writing analysis results to Excel files
- Data validation and formatting
"""

import logging
from pathlib import Path
from typing import List, Tuple, Optional
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


@dataclass
class RepositoryEntry:
    """Single repository entry from input Excel file."""
    row_index: int
    data: dict  # Original row data from Excel
    url: Optional[str] = None  # Extracted GitHub URL

    def __post_init__(self):
        """Extract URL from data if not provided."""
        if self.url is None and self.data:
            # Try to find URL in common column names
            for key in ['Repo URL', 'url', 'URL', 'github_url', 'GitHub URL', 'repository', 'Repository']:
                if key in self.data:
                    self.url = self.data[key]
                    break


class ExcelReader:
    """Reads repository data from Excel files."""

    @staticmethod
    def read_input_file(file_path: Path) -> List[RepositoryEntry]:
        """
        Read repository entries from input Excel file.

        Args:
            file_path: Path to input Excel file (output/output12.xlsx)

        Returns:
            List of RepositoryEntry objects

        Raises:
            FileNotFoundError: If file does not exist
            Exception: If file cannot be read
        """
        if not file_path.exists():
            logger.error(f"Input file not found: {file_path}")
            raise FileNotFoundError(f"Input file not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            if worksheet is None:
                logger.error(f"No active worksheet in {file_path}")
                raise ValueError("No active worksheet found")

            entries = []
            headers = []
            column_count = 0

            # Read header row
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value is None:
                    break
                headers.append(cell.value)
                column_count = col_idx

            if not headers:
                logger.warning(f"No headers found in {file_path}")
                return []

            logger.info(f"Found headers: {headers}")

            # Read data rows
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                row_data = {}
                has_data = False

                for col_idx, (header, cell) in enumerate(zip(headers, row[:column_count]), 1):
                    value = cell.value
                    row_data[header] = value
                    if value is not None:
                        has_data = True

                if has_data:
                    entry = RepositoryEntry(
                        row_index=row_idx,
                        data=row_data
                    )
                    entries.append(entry)

            logger.info(f"Read {len(entries)} entries from {file_path}")
            workbook.close()
            return entries

        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            raise

    @staticmethod
    def extract_urls(entries: List[RepositoryEntry]) -> List[RepositoryEntry]:
        """
        Extract and validate GitHub URLs from entries.

        Args:
            entries: List of repository entries

        Returns:
            List of entries with extracted URLs
        """
        for entry in entries:
            if entry.url:
                # Ensure URL is a string
                entry.url = str(entry.url).strip()
                if not entry.url.startswith(('http://', 'https://', 'git@')):
                    # Try to make it a valid URL if it's missing scheme
                    if entry.url.startswith('github.com'):
                        entry.url = f"https://{entry.url}"

        entries_with_urls = [e for e in entries if e.url]
        logger.info(f"Extracted {len(entries_with_urls)} entries with URLs from {len(entries)} total")

        return entries_with_urls


class ExcelWriter:
    """Writes analysis results to Excel files."""

    GRADE_COLUMN_NAME = "grade"
    GRADE_COLUMN_WIDTH = 12
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    GRADE_DECIMAL_PLACES = 2

    @staticmethod
    def write_output_file(
        file_path: Path,
        entries: List[RepositoryEntry],
        grades: dict  # URL -> grade value
    ) -> bool:
        """
        Write analysis results to output Excel file.

        Args:
            file_path: Path to output Excel file (output/output23.xlsx)
            entries: Original repository entries
            grades: Dictionary mapping URLs to grade values

        Returns:
            True if successful, False otherwise
        """
        try:
            # Create new workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Analysis Results"

            if not entries or not entries[0].data:
                logger.warning("No entries to write")
                return False

            # Get headers from first entry
            headers = list(entries[0].data.keys())
            headers.append(ExcelWriter.GRADE_COLUMN_NAME)

            # Write header row
            for col_idx, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = ExcelWriter.HEADER_FILL
                cell.font = ExcelWriter.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows
            for row_idx, entry in enumerate(entries, start=2):
                # Write original columns
                for col_idx, header in enumerate(headers[:-1], 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = entry.data.get(header)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Write grade column
                grade = grades.get(entry.url, 0.0)
                grade_cell = worksheet.cell(row=row_idx, column=len(headers))
                grade_cell.value = round(float(grade), ExcelWriter.GRADE_DECIMAL_PLACES)
                grade_cell.alignment = Alignment(horizontal="center", vertical="center")
                grade_cell.number_format = f"0.{'0' * ExcelWriter.GRADE_DECIMAL_PLACES}"

            # Adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

            # Ensure output directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            workbook.save(file_path)
            logger.info(f"Wrote {len(entries)} entries to {file_path}")
            workbook.close()
            return True

        except Exception as e:
            logger.error(f"Error writing Excel file {file_path}: {e}")
            return False

"""
Unit tests for the excel_processor module.

Tests cover:
- Reading Excel files
- Writing Excel files
- URL extraction
- Error handling
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import openpyxl

from excel_processor import (
    RepositoryEntry,
    ExcelReader,
    ExcelWriter
)


class TestRepositoryEntry:
    """Tests for RepositoryEntry dataclass."""

    def test_init_with_url(self):
        """Test initialization with provided URL."""
        entry = RepositoryEntry(
            row_index=2,
            data={"name": "repo", "url": "https://github.com/test/repo"},
            url="https://github.com/test/repo"
        )

        assert entry.row_index == 2
        assert entry.url == "https://github.com/test/repo"

    def test_init_extract_url_from_data(self):
        """Test URL extraction from data dictionary."""
        entry = RepositoryEntry(
            row_index=2,
            data={"name": "repo", "url": "https://github.com/test/repo"}
        )

        assert entry.url == "https://github.com/test/repo"

    def test_init_extract_url_case_insensitive(self):
        """Test URL extraction with different case variants."""
        entry = RepositoryEntry(
            row_index=2,
            data={"name": "repo", "URL": "https://github.com/test/repo"}
        )

        assert entry.url == "https://github.com/test/repo"

    def test_init_no_url(self):
        """Test initialization with no URL."""
        entry = RepositoryEntry(
            row_index=2,
            data={"name": "repo"}
        )

        assert entry.url is None

    def test_init_empty_data(self):
        """Test initialization with empty data."""
        entry = RepositoryEntry(
            row_index=2,
            data={}
        )

        assert entry.url is None

    def test_init_extract_url_from_repo_url_field(self):
        """Test URL extraction from 'Repo URL' field (Agent 1 output format)."""
        entry = RepositoryEntry(
            row_index=2,
            data={"email": "test@example.com", "Repo URL": "https://github.com/test/repo"}
        )

        assert entry.url == "https://github.com/test/repo"


class TestExcelReader:
    """Tests for ExcelReader class."""

    def test_read_input_file_basic(self, tmp_path):
        """Test reading a basic Excel file."""
        # Create test Excel file
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "name"
        ws['B1'] = "url"
        ws['A2'] = "repo1"
        ws['B2'] = "https://github.com/test/repo1"
        wb.save(excel_file)

        entries = ExcelReader.read_input_file(excel_file)

        assert len(entries) == 1
        assert entries[0].row_index == 2
        assert entries[0].data["name"] == "repo1"
        assert entries[0].data["url"] == "https://github.com/test/repo1"

    def test_read_input_file_multiple_rows(self, tmp_path):
        """Test reading Excel file with multiple rows."""
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "name"
        ws['B1'] = "url"
        ws['A2'] = "repo1"
        ws['B2'] = "https://github.com/test/repo1"
        ws['A3'] = "repo2"
        ws['B3'] = "https://github.com/test/repo2"
        wb.save(excel_file)

        entries = ExcelReader.read_input_file(excel_file)

        assert len(entries) == 2
        assert entries[0].data["name"] == "repo1"
        assert entries[1].data["name"] == "repo2"

    def test_read_input_file_not_found(self, tmp_path):
        """Test error handling for nonexistent file."""
        excel_file = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            ExcelReader.read_input_file(excel_file)

    def test_read_input_file_empty(self, tmp_path):
        """Test reading empty Excel file."""
        excel_file = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "header"
        wb.save(excel_file)

        entries = ExcelReader.read_input_file(excel_file)

        assert len(entries) == 0

    def test_read_input_file_multiple_columns(self, tmp_path):
        """Test reading Excel file with multiple columns."""
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "name"
        ws['B1'] = "url"
        ws['C1'] = "description"
        ws['A2'] = "repo1"
        ws['B2'] = "https://github.com/test/repo1"
        ws['C2'] = "Test repository"
        wb.save(excel_file)

        entries = ExcelReader.read_input_file(excel_file)

        assert len(entries) == 1
        assert entries[0].data["name"] == "repo1"
        assert entries[0].data["url"] == "https://github.com/test/repo1"
        assert entries[0].data["description"] == "Test repository"

    def test_extract_urls_valid(self):
        """Test extracting valid URLs."""
        entries = [
            RepositoryEntry(2, {"url": "https://github.com/test/repo1"}),
            RepositoryEntry(3, {"url": "https://github.com/test/repo2"}),
        ]

        result = ExcelReader.extract_urls(entries)

        assert len(result) == 2
        assert all(e.url for e in result)

    def test_extract_urls_no_urls(self):
        """Test extracting URLs when none exist."""
        entries = [
            RepositoryEntry(2, {"name": "repo1"}),
            RepositoryEntry(3, {"name": "repo2"}),
        ]

        result = ExcelReader.extract_urls(entries)

        assert len(result) == 0

    def test_extract_urls_mixed(self):
        """Test extracting URLs with mixed entries."""
        entries = [
            RepositoryEntry(2, {"url": "https://github.com/test/repo1"}),
            RepositoryEntry(3, {"name": "repo2"}),
        ]

        result = ExcelReader.extract_urls(entries)

        assert len(result) == 1
        assert result[0].url == "https://github.com/test/repo1"

    def test_extract_urls_adds_scheme(self):
        """Test URL extraction adds https scheme if missing."""
        entries = [
            RepositoryEntry(2, {"url": "github.com/test/repo"}),
        ]

        result = ExcelReader.extract_urls(entries)

        assert len(result) == 1
        assert result[0].url == "https://github.com/test/repo"

    def test_extract_urls_whitespace(self):
        """Test URL extraction strips whitespace."""
        entries = [
            RepositoryEntry(2, {"url": "  https://github.com/test/repo  "}),
        ]

        result = ExcelReader.extract_urls(entries)

        assert len(result) == 1
        assert result[0].url == "https://github.com/test/repo"


class TestExcelWriter:
    """Tests for ExcelWriter class."""

    def test_write_output_file_basic(self, tmp_path):
        """Test writing basic output file."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {
                "name": "repo1",
                "url": "https://github.com/test/repo1"
            }),
        ]
        grades = {"https://github.com/test/repo1": 75.5}

        success = ExcelWriter.write_output_file(output_file, entries, grades)

        assert success is True
        assert output_file.exists()

        # Verify file contents
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws['A1'].value == "name"
        assert ws['B1'].value == "url"
        assert ws['C1'].value == "grade"
        assert ws['A2'].value == "repo1"
        assert ws['C2'].value == 75.5

    def test_write_output_file_multiple_entries(self, tmp_path):
        """Test writing output file with multiple entries."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {
                "name": "repo1",
                "url": "https://github.com/test/repo1"
            }),
            RepositoryEntry(3, {
                "name": "repo2",
                "url": "https://github.com/test/repo2"
            }),
        ]
        grades = {
            "https://github.com/test/repo1": 75.5,
            "https://github.com/test/repo2": 88.0,
        }

        success = ExcelWriter.write_output_file(output_file, entries, grades)

        assert success is True

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws['A3'].value == "repo2"
        assert ws['C3'].value == 88.0

    def test_write_output_file_creates_directory(self, tmp_path):
        """Test that write creates necessary directories."""
        output_file = tmp_path / "subdir" / "output.xlsx"
        entries = [
            RepositoryEntry(2, {"name": "repo1"}),
        ]

        success = ExcelWriter.write_output_file(output_file, entries, {})

        assert success is True
        assert output_file.exists()

    def test_write_output_file_empty_entries(self, tmp_path):
        """Test error handling with empty entries."""
        output_file = tmp_path / "output.xlsx"

        success = ExcelWriter.write_output_file(output_file, [], {})

        assert success is False

    def test_write_output_file_grade_formatting(self, tmp_path):
        """Test that grades are formatted to 2 decimal places."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {"name": "repo1", "url": "https://github.com/test/repo1"}),
        ]
        grades = {"https://github.com/test/repo1": 75.123456}

        ExcelWriter.write_output_file(output_file, entries, grades)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        grade_value = ws['C2'].value
        assert grade_value == 75.12

    def test_write_output_file_preserves_original_data(self, tmp_path):
        """Test that original data is preserved in output."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {
                "name": "repo1",
                "url": "https://github.com/test/repo1",
                "description": "A test repo"
            }),
        ]

        ExcelWriter.write_output_file(output_file, entries, {})

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws['A1'].value == "name"
        assert ws['B1'].value == "url"
        assert ws['C1'].value == "description"
        assert ws['C2'].value == "A test repo"

    def test_write_output_file_missing_grade(self, tmp_path):
        """Test handling of missing grades."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {"name": "repo1", "url": "https://github.com/test/repo1"}),
        ]
        grades = {}  # No grade for this URL

        ExcelWriter.write_output_file(output_file, entries, grades)

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws['C2'].value == 0.0

    def test_write_output_file_header_formatting(self, tmp_path):
        """Test that headers are properly formatted."""
        output_file = tmp_path / "output.xlsx"
        entries = [
            RepositoryEntry(2, {"name": "repo1"}),
        ]

        ExcelWriter.write_output_file(output_file, entries, {})

        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        header_cell = ws['A1']
        assert header_cell.value == "name"
        assert header_cell.font.bold is True
        # Color format includes alpha channel: 00FFFFFF (white)
        assert header_cell.font.color.rgb in ("FFFFFFFF", "00FFFFFF")

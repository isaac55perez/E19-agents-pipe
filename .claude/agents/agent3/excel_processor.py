"""
Excel file reading and writing functionality for grade data transformation.

This module handles:
- Reading input Excel files with grade data
- Writing output Excel files with greeting messages
- Data validation and error handling
"""

import logging
from pathlib import Path
from typing import List, Dict, Optional, Tuple

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

logger = logging.getLogger(__name__)


class ExcelReader:
    """Read Excel files containing grade data."""

    @staticmethod
    def read_input_file(file_path: Path) -> Tuple[List[Dict], List[str]]:
        """
        Read input Excel file with grade data.

        Args:
            file_path: Path to input Excel file

        Returns:
            Tuple of (rows_data, headers) where rows_data is list of dicts

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If file is invalid or missing required columns
        """
        if not file_path.exists():
            logger.error(f"Input file not found: {file_path}")
            raise FileNotFoundError(f"Input file not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            if worksheet is None:
                raise ValueError("No active worksheet found")

            # Read headers
            headers = []
            for cell in worksheet[1]:
                if cell.value is None:
                    break
                headers.append(cell.value)

            if not headers:
                logger.warning("No headers found in input file")
                return [], []

            logger.info(f"Found headers: {headers}")

            # Verify grade column exists
            if "grade" not in headers:
                raise ValueError("'grade' column not found in input file")

            # Read data rows
            rows_data = []
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=False), start=2):
                row_data = {}
                has_data = False

                for header, cell in zip(headers, row[: len(headers)]):
                    value = cell.value
                    row_data[header] = value
                    if value is not None:
                        has_data = True

                if has_data:
                    rows_data.append(row_data)

            logger.info(f"Read {len(rows_data)} data rows from {file_path}")
            workbook.close()
            return rows_data, headers

        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            raise

    @staticmethod
    def validate_grades(rows_data: List[Dict]) -> Tuple[bool, List[str]]:
        """
        Validate grade values in data rows.

        Args:
            rows_data: List of row dictionaries

        Returns:
            Tuple of (is_valid, error_messages)
        """
        errors = []

        for idx, row in enumerate(rows_data, start=2):
            grade = row.get("grade")

            if grade is None:
                errors.append(f"Row {idx}: Missing grade value")
                continue

            try:
                grade_float = float(grade)
                if not 0 <= grade_float <= 100:
                    errors.append(f"Row {idx}: Grade out of range (0-100): {grade_float}")
            except (ValueError, TypeError):
                errors.append(f"Row {idx}: Grade is not numeric: {grade}")

        if errors:
            logger.warning(f"Found {len(errors)} grade validation errors")
            for error in errors:
                logger.warning(f"  {error}")

        return len(errors) == 0, errors


class ExcelWriter:
    """Write Excel files with greeting messages."""

    GREETING_COLUMN_NAME = "greeting"
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    @staticmethod
    def validate_grades(rows_data: List[Dict]) -> Tuple[bool, List[str]]:
        """
        Validate grade values in data rows.

        Args:
            rows_data: List of row dictionaries

        Returns:
            Tuple of (is_valid, error_messages)
        """
        return ExcelReader.validate_grades(rows_data)

    @staticmethod
    def write_output_file(
        file_path: Path,
        headers: List[str],
        rows_data: List[Dict],
        greetings: Dict[int, str],
    ) -> bool:
        """
        Write output Excel file with greeting messages.

        Args:
            file_path: Path to output Excel file
            headers: Original column headers
            rows_data: Original row data
            greetings: Dict mapping row index to greeting message

        Returns:
            True if successful, False otherwise
        """
        try:
            # Create new workbook
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Results"

            if not headers or not rows_data:
                logger.warning("No data to write")
                return False

            # Create new headers with greeting column
            new_headers = headers + [ExcelWriter.GREETING_COLUMN_NAME]

            # Write header row
            for col_idx, header in enumerate(new_headers, 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = ExcelWriter.HEADER_FILL
                cell.font = ExcelWriter.HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Write data rows
            for row_idx, row_data in enumerate(rows_data, start=2):
                # Write original columns
                for col_idx, header in enumerate(headers, 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = row_data.get(header)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Write greeting column
                greeting = greetings.get(row_idx - 2, "")
                greeting_cell = worksheet.cell(row=row_idx, column=len(new_headers))
                greeting_cell.value = greeting
                greeting_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Adjust column widths
            for col_idx, header in enumerate(new_headers, 1):
                max_length = len(str(header))
                for row in worksheet.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 50)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = adjusted_width

            # Ensure output directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Save workbook
            workbook.save(file_path)
            logger.info(f"Wrote {len(rows_data)} entries to {file_path}")
            workbook.close()
            return True

        except Exception as e:
            logger.error(f"Error writing Excel file {file_path}: {e}")
            return False

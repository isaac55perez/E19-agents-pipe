"""
Unit tests for excel_processor module.

Tests cover:
- Reading Excel files
- Writing Excel files
- Data validation
- Error handling
"""

import pytest
from pathlib import Path
import openpyxl

from excel_processor import ExcelReader, ExcelWriter


class TestExcelReader:
    """Tests for ExcelReader class."""

    def test_read_input_file_basic(self, tmp_path):
        """Test reading a basic Excel file."""
        # Create test Excel file
        excel_file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "grade"])
        ws.append([1, "Alice", 85.0])
        ws.append([2, "Bob", 45.0])
        wb.save(excel_file)

        # Read file
        rows_data, headers = ExcelReader.read_input_file(excel_file)

        assert len(headers) == 3
        assert headers == ["ID", "Name", "grade"]
        assert len(rows_data) == 2
        assert rows_data[0]["ID"] == 1
        assert rows_data[0]["grade"] == 85.0

    def test_read_input_file_not_found(self):
        """Test reading non-existent file."""
        with pytest.raises(FileNotFoundError):
            ExcelReader.read_input_file(Path("nonexistent.xlsx"))

    def test_read_input_file_no_grade_column(self, tmp_path):
        """Test reading file without grade column."""
        excel_file = tmp_path / "no_grade.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name"])
        ws.append([1, "Alice"])
        wb.save(excel_file)

        with pytest.raises(ValueError):
            ExcelReader.read_input_file(excel_file)

    def test_read_input_file_multiple_rows(self, tmp_path):
        """Test reading file with multiple rows."""
        excel_file = tmp_path / "multiple.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "grade"])
        for i in range(1, 11):
            ws.append([i, float(i * 10)])
        wb.save(excel_file)

        rows_data, headers = ExcelReader.read_input_file(excel_file)
        assert len(rows_data) == 10
        assert rows_data[0]["grade"] == 10.0
        assert rows_data[9]["grade"] == 100.0

    def test_read_input_file_empty(self, tmp_path):
        """Test reading empty file (no data rows)."""
        excel_file = tmp_path / "empty.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "grade"])
        wb.save(excel_file)

        rows_data, headers = ExcelReader.read_input_file(excel_file)
        assert len(rows_data) == 0
        assert len(headers) == 2

    def test_read_input_file_mixed_types(self, tmp_path):
        """Test reading file with mixed data types."""
        excel_file = tmp_path / "mixed.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "URL", "grade"])
        ws.append([1, "Alice", "https://github.com/alice", 85.5])
        wb.save(excel_file)

        rows_data, headers = ExcelReader.read_input_file(excel_file)
        assert len(rows_data) == 1
        assert rows_data[0]["Name"] == "Alice"
        assert rows_data[0]["URL"] == "https://github.com/alice"

    def test_validate_grades_valid(self):
        """Test grade validation with valid grades."""
        rows_data = [
            {"grade": 85.0},
            {"grade": 45.0},
            {"grade": 100.0},
        ]
        is_valid, errors = ExcelWriter.validate_grades(rows_data)
        assert is_valid
        assert len(errors) == 0

    def test_validate_grades_out_of_range(self):
        """Test grade validation with out-of-range grades."""
        rows_data = [
            {"grade": 85.0},
            {"grade": 105.0},  # Out of range
        ]
        is_valid, errors = ExcelWriter.validate_grades(rows_data)
        assert not is_valid
        assert len(errors) > 0

    def test_validate_grades_non_numeric(self):
        """Test grade validation with non-numeric grades."""
        rows_data = [
            {"grade": "invalid"},
        ]
        is_valid, errors = ExcelWriter.validate_grades(rows_data)
        assert not is_valid
        assert len(errors) > 0

    def test_validate_grades_missing(self):
        """Test grade validation with missing grades."""
        rows_data = [
            {"grade": None},
        ]
        is_valid, errors = ExcelWriter.validate_grades(rows_data)
        assert not is_valid
        assert len(errors) > 0


class TestExcelWriter:
    """Tests for ExcelWriter class."""

    def test_write_output_file_basic(self, tmp_path):
        """Test writing basic output file."""
        output_file = tmp_path / "output.xlsx"
        headers = ["ID", "Name", "grade"]
        rows_data = [
            {"ID": 1, "Name": "Alice", "grade": 85.0},
            {"ID": 2, "Name": "Bob", "grade": 45.0},
        ]
        greetings = {0: "Great job!", 1: "Keep trying!"}

        success = ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        assert success
        assert output_file.exists()

        # Verify output
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws.cell(1, 1).value == "ID"
        assert ws.cell(1, 4).value == "greeting"
        assert ws.cell(2, 4).value == "Great job!"
        assert ws.cell(3, 4).value == "Keep trying!"

    def test_write_output_file_creates_directory(self, tmp_path):
        """Test that write_output_file creates missing directories."""
        output_file = tmp_path / "subdir" / "deep" / "output.xlsx"
        headers = ["ID", "grade"]
        rows_data = [{"ID": 1, "grade": 85.0}]
        greetings = {0: "Great!"}

        success = ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        assert success
        assert output_file.exists()

    def test_write_output_file_empty_data(self, tmp_path):
        """Test writing file with no data."""
        output_file = tmp_path / "empty.xlsx"
        headers = ["ID", "grade"]
        rows_data = []
        greetings = {}

        success = ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        assert not success

    def test_write_output_file_preserves_data(self, tmp_path):
        """Test that output file preserves original data."""
        output_file = tmp_path / "output.xlsx"
        headers = ["ID", "Name", "grade"]
        rows_data = [
            {"ID": 1, "Name": "Alice", "grade": 85.0},
        ]
        greetings = {0: "Great!"}

        ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        # Read back
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Alice"
        assert ws.cell(2, 3).value == 85.0
        assert ws.cell(2, 4).value == "Great!"

    def test_write_output_file_formatting(self, tmp_path):
        """Test that output file has proper formatting."""
        output_file = tmp_path / "output.xlsx"
        headers = ["ID", "grade"]
        rows_data = [{"ID": 1, "grade": 85.0}]
        greetings = {0: "Great!"}

        ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        # Check formatting
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        header_cell = ws.cell(1, 1)
        # Header should have fill
        assert header_cell.fill is not None

    def test_write_output_file_multiple_entries(self, tmp_path):
        """Test writing file with multiple entries."""
        output_file = tmp_path / "output.xlsx"
        headers = ["ID", "grade"]
        rows_data = [{"ID": i, "grade": float(i * 10)} for i in range(1, 6)]
        greetings = {i: f"Message {i}" for i in range(5)}

        success = ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        assert success
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active
        # Check we have correct number of rows (header + 5 data)
        assert ws.max_row == 6

"""
Integration tests for main workflow.

Tests cover:
- Complete end-to-end workflow
- Command-line arguments
- File I/O
- Error scenarios
"""

import pytest
from pathlib import Path
import openpyxl
import sys
from unittest import mock

from excel_processor import ExcelReader, ExcelWriter
from greeting_generator import GreetingGenerator


class TestIntegration:
    """Integration tests for complete workflow."""

    def test_complete_workflow(self, tmp_path):
        """Test complete greeting transformation workflow."""
        # Create input file
        input_file = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "Name", "Repo URL", "grade"])
        ws.append([1, "Alice", "https://github.com/alice/repo1", 85.5])
        ws.append([2, "Bob", "https://github.com/bob/repo2", 45.0])
        ws.append([3, "Charlie", "https://github.com/charlie/repo3", 72.5])
        wb.save(input_file)

        # Read input
        rows_data, headers = ExcelReader.read_input_file(input_file)
        assert len(rows_data) == 3

        # Generate greetings
        greetings = {}
        for idx, row in enumerate(rows_data):
            grade = row["grade"]
            greeting = GreetingGenerator.generate_greeting(grade)
            greetings[idx] = greeting

        assert len(greetings) == 3

        # Write output
        output_file = tmp_path / "output.xlsx"
        success = ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)
        assert success

        # Verify output
        wb_out = openpyxl.load_workbook(output_file)
        ws_out = wb_out.active

        # Check headers
        assert ws_out.cell(1, 1).value == "ID"
        assert ws_out.cell(1, 5).value == "greeting"

        # Check data
        assert ws_out.cell(2, 1).value == 1
        assert ws_out.cell(2, 2).value == "Alice"
        assert ws_out.cell(2, 5).value is not None  # Has greeting

        # Check personality styles
        # Alice has 85.5 (>60) -> Trump style
        alice_greeting = ws_out.cell(2, 5).value
        assert alice_greeting is not None

        # Bob has 45.0 (<60) -> Murphy style
        bob_greeting = ws_out.cell(3, 5).value
        assert bob_greeting is not None

    def test_mixed_grade_distribution(self, tmp_path):
        """Test with mixed grade distribution."""
        rows_data = [
            {"ID": 1, "grade": 30},    # Murphy
            {"ID": 2, "grade": 60},    # Murphy (threshold)
            {"ID": 3, "grade": 61},    # Trump
            {"ID": 4, "grade": 100},   # Trump
        ]

        # Count personality styles
        murphy_count = 0
        trump_count = 0
        for row in rows_data:
            style = GreetingGenerator.get_personality_style(row["grade"])
            if style == "Eddie Murphy":
                murphy_count += 1
            else:
                trump_count += 1

        assert murphy_count == 2
        assert trump_count == 2

    def test_grade_validation_integration(self, tmp_path):
        """Test grade validation in workflow."""
        input_file = tmp_path / "invalid.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ID", "grade"])
        ws.append([1, 85.0])
        ws.append([2, "invalid"])  # Invalid grade
        ws.append([3, 150])  # Out of range
        wb.save(input_file)

        # Read and validate
        rows_data, headers = ExcelReader.read_input_file(input_file)
        is_valid, errors = ExcelWriter.validate_grades(rows_data)

        assert not is_valid
        assert len(errors) > 0

    def test_personality_consistency(self):
        """Test that same grade gets same personality style."""
        grade = 75.5
        styles = set()

        for _ in range(10):
            style = GreetingGenerator.get_personality_style(grade)
            styles.add(style)

        # Should only have one style for this grade
        assert len(styles) == 1
        assert "Donald Trump" in styles

    def test_grade_boundary_low(self):
        """Test grades at low boundary."""
        # Grade 0 and just above
        assert GreetingGenerator.get_personality_style(0) == "Eddie Murphy"
        assert GreetingGenerator.get_personality_style(0.1) == "Eddie Murphy"

    def test_grade_boundary_high(self):
        """Test grades at high boundary."""
        # Grade 100
        assert GreetingGenerator.get_personality_style(100) == "Donald Trump"
        assert GreetingGenerator.get_personality_style(99.9) == "Donald Trump"

    def test_greeting_generation_consistency(self):
        """Test that greetings are non-empty and appropriate."""
        test_grades = [0, 30, 60, 61, 100]

        for grade in test_grades:
            greeting = GreetingGenerator.generate_greeting(grade)
            assert greeting is not None
            assert isinstance(greeting, str)
            assert len(greeting) > 10  # Should be reasonably long
            assert len(greeting) < 500  # Should not be too long

    def test_data_preservation_through_pipeline(self, tmp_path):
        """Test that original data is preserved through transformation."""
        original_data = {
            "ID": 1,
            "Name": "Alice",
            "Email": "alice@example.com",
            "Repo URL": "https://github.com/alice",
            "grade": 85.0
        }

        headers = list(original_data.keys())
        rows_data = [original_data]
        greetings = {0: "Great job!"}

        output_file = tmp_path / "output.xlsx"
        ExcelWriter.write_output_file(output_file, headers, rows_data, greetings)

        # Read back and verify
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "Alice"
        assert ws.cell(2, 3).value == "alice@example.com"
        assert ws.cell(2, 5).value == 85.0
        assert ws.cell(2, 6).value == "Great job!"

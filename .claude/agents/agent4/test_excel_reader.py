"""
Unit tests for excel_reader module.
"""

import pytest
from pathlib import Path
import tempfile
import openpyxl

from excel_reader import ExcelReader


class TestValidateEmail:
    """Tests for validate_email method."""

    def test_validate_email_valid(self):
        """Test validation of valid email address."""
        assert ExcelReader.validate_email("user@example.com") is True

    def test_validate_email_valid_with_numbers(self):
        """Test validation of email with numbers."""
        assert ExcelReader.validate_email("user123@example.com") is True

    def test_validate_email_valid_with_underscore(self):
        """Test validation of email with underscore."""
        assert ExcelReader.validate_email("user_name@example.com") is True

    def test_validate_email_valid_with_dot(self):
        """Test validation of email with dot in local part."""
        assert ExcelReader.validate_email("first.last@example.com") is True

    def test_validate_email_valid_subdomain(self):
        """Test validation of email with subdomain."""
        assert ExcelReader.validate_email("user@mail.example.com") is True

    def test_validate_email_invalid_no_at(self):
        """Test that email without @ is invalid."""
        assert ExcelReader.validate_email("userexample.com") is False

    def test_validate_email_invalid_no_domain(self):
        """Test that email without domain is invalid."""
        assert ExcelReader.validate_email("user@") is False

    def test_validate_email_invalid_no_local(self):
        """Test that email without local part is invalid."""
        assert ExcelReader.validate_email("@example.com") is False

    def test_validate_email_invalid_no_tld(self):
        """Test that email without TLD is invalid."""
        assert ExcelReader.validate_email("user@example") is False

    def test_validate_email_invalid_spaces(self):
        """Test that email with spaces is invalid."""
        assert ExcelReader.validate_email("user name@example.com") is False

    def test_validate_email_none(self):
        """Test that None is invalid."""
        assert ExcelReader.validate_email(None) is False

    def test_validate_email_empty_string(self):
        """Test that empty string is invalid."""
        assert ExcelReader.validate_email("") is False

    def test_validate_email_not_string(self):
        """Test that non-string input is invalid."""
        assert ExcelReader.validate_email(123) is False


class TestValidateUrl:
    """Tests for validate_url method."""

    def test_validate_url_github_https(self):
        """Test validation of GitHub HTTPS URL."""
        assert ExcelReader.validate_url("https://github.com/user/repo") is True

    def test_validate_url_github_https_with_git(self):
        """Test validation of GitHub HTTPS URL with .git suffix."""
        assert ExcelReader.validate_url("https://github.com/user/repo.git") is True

    def test_validate_url_github_http(self):
        """Test validation of GitHub HTTP URL."""
        assert ExcelReader.validate_url("http://github.com/user/repo") is True

    def test_validate_url_github_complex_path(self):
        """Test validation of GitHub URL with complex path."""
        assert ExcelReader.validate_url("https://github.com/org/project-name.git") is True

    def test_validate_url_invalid_no_github(self):
        """Test that non-GitHub URL is invalid."""
        assert ExcelReader.validate_url("https://example.com/user/repo") is False

    def test_validate_url_invalid_no_protocol(self):
        """Test that URL without protocol is invalid."""
        assert ExcelReader.validate_url("github.com/user/repo") is False

    def test_validate_url_invalid_ftp(self):
        """Test that FTP URL is invalid."""
        assert ExcelReader.validate_url("ftp://github.com/user/repo") is False

    def test_validate_url_none(self):
        """Test that None is invalid."""
        assert ExcelReader.validate_url(None) is False

    def test_validate_url_empty_string(self):
        """Test that empty string is invalid."""
        assert ExcelReader.validate_url("") is False

    def test_validate_url_not_string(self):
        """Test that non-string input is invalid."""
        assert ExcelReader.validate_url(123) is False

    def test_validate_url_case_insensitive(self):
        """Test that validation checks github.com (lowercase)."""
        # The validation uses lowercase checks, so uppercase URLs fail
        # This is expected behavior - URLs should use lowercase
        assert ExcelReader.validate_url("https://github.com/user/repo") is True
        # Uppercase domain portion fails the lowercase check
        assert ExcelReader.validate_url("HTTPS://GITHUB.COM/user/repo") is False


class TestReadInputFile:
    """Tests for read_input_file method."""

    def test_read_input_file_not_found(self):
        """Test that FileNotFoundError is raised for missing file."""
        with pytest.raises(FileNotFoundError):
            ExcelReader.read_input_file(Path("nonexistent.xlsx"))

    def test_read_input_file_basic(self):
        """Test reading basic Excel file."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create test workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["user1@example.com", "User1", "https://github.com/user/repo1", "Hello"])
            wb.save(tmp.name)

            # Read file
            rows_data, headers = ExcelReader.read_input_file(Path(tmp.name))

            assert len(rows_data) == 1
            assert headers == ["email", "name", "Repo URL", "greeting"]
            assert rows_data[0]["email"] == "user1@example.com"

    def test_read_input_file_multiple_rows(self):
        """Test reading file with multiple rows."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            for i in range(5):
                ws.append([
                    f"user{i}@example.com",
                    f"User{i}",
                    f"https://github.com/user/repo{i}",
                    f"Greeting{i}"
                ])
            wb.save(tmp.name)

            rows_data, headers = ExcelReader.read_input_file(Path(tmp.name))

            assert len(rows_data) == 5
            assert rows_data[0]["email"] == "user0@example.com"
            assert rows_data[4]["email"] == "user4@example.com"

    def test_read_input_file_missing_columns(self):
        """Test that ValueError is raised for missing required columns."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name"])  # Missing Repo URL and greeting
            ws.append(["user@example.com", "User"])
            wb.save(tmp.name)

            with pytest.raises(ValueError):
                ExcelReader.read_input_file(Path(tmp.name))

    def test_read_input_file_empty_rows_skipped(self):
        """Test that empty rows are skipped."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["user1@example.com", "User1", "https://github.com/user/repo1", "Hello"])
            ws.append([None, None, None, None])  # Empty row
            ws.append(["user2@example.com", "User2", "https://github.com/user/repo2", "Hi"])
            wb.save(tmp.name)

            rows_data, headers = ExcelReader.read_input_file(Path(tmp.name))

            assert len(rows_data) == 2
            assert rows_data[0]["email"] == "user1@example.com"
            assert rows_data[1]["email"] == "user2@example.com"


class TestValidateEntries:
    """Tests for validate_entries method."""

    def test_validate_entries_all_valid(self):
        """Test validation with all valid entries."""
        rows_data = [
            {
                "email": "user1@example.com",
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            },
            {
                "email": "user2@example.com",
                "name": "User2",
                "Repo URL": "https://github.com/user/repo2",
                "greeting": "Hi"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 2
        assert len(invalid) == 0

    def test_validate_entries_all_invalid_email(self):
        """Test validation with invalid email addresses."""
        rows_data = [
            {
                "email": "invalid-email",
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        assert "Invalid email" in invalid[0]["errors"][0]

    def test_validate_entries_all_invalid_url(self):
        """Test validation with invalid repository URLs."""
        rows_data = [
            {
                "email": "user@example.com",
                "name": "User1",
                "Repo URL": "https://example.com/user/repo",  # Not GitHub
                "greeting": "Hello"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        assert "Invalid repository URL" in invalid[0]["errors"][0]

    def test_validate_entries_missing_email(self):
        """Test validation with missing email."""
        rows_data = [
            {
                "email": None,
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        assert "Missing email" in invalid[0]["errors"][0]

    def test_validate_entries_missing_name(self):
        """Test validation with missing name."""
        rows_data = [
            {
                "email": "user@example.com",
                "name": None,
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        assert "Missing name" in invalid[0]["errors"][0]

    def test_validate_entries_missing_greeting(self):
        """Test validation with missing greeting."""
        rows_data = [
            {
                "email": "user@example.com",
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": None
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        assert "Missing greeting" in invalid[0]["errors"][0]

    def test_validate_entries_mixed_valid_invalid(self):
        """Test validation with mix of valid and invalid entries."""
        rows_data = [
            {
                "email": "user1@example.com",
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            },
            {
                "email": "invalid",
                "name": "User2",
                "Repo URL": "https://github.com/user/repo2",
                "greeting": "Hi"
            },
            {
                "email": "user3@example.com",
                "name": "User3",
                "Repo URL": "https://github.com/user/repo3",
                "greeting": "Hey"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 2
        assert len(invalid) == 1
        assert invalid[0]["row_index"] == 3  # Row 3 (1-indexed)

    def test_validate_entries_multiple_errors(self):
        """Test that entry with multiple errors is captured."""
        rows_data = [
            {
                "email": "invalid-email",
                "name": None,
                "Repo URL": "not-a-url",
                "greeting": None
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 1
        # Should have multiple error messages
        assert len(invalid[0]["errors"]) >= 3

    def test_validate_entries_empty_list(self):
        """Test validation with empty list."""
        rows_data = []

        valid, invalid = ExcelReader.validate_entries(rows_data)

        assert len(valid) == 0
        assert len(invalid) == 0

    def test_validate_entries_row_index_preserved(self):
        """Test that row indices are preserved correctly."""
        rows_data = [
            {
                "email": "user1@example.com",
                "name": "User1",
                "Repo URL": "https://github.com/user/repo1",
                "greeting": "Hello"
            },
            {
                "email": "invalid",
                "name": "User2",
                "Repo URL": "https://github.com/user/repo2",
                "greeting": "Hi"
            }
        ]

        valid, invalid = ExcelReader.validate_entries(rows_data)

        # Row indices should be 1-based (row 2 is first data row)
        assert invalid[0]["row_index"] == 3

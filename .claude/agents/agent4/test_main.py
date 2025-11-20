"""
Integration tests for main module.
"""

import pytest
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock
import openpyxl

from main import run_workflow, parse_arguments


class TestRunWorkflow:
    """Tests for run_workflow function."""

    def test_run_workflow_input_file_not_found(self):
        """Test workflow with non-existent input file."""
        results = run_workflow(
            input_file="nonexistent.xlsx",
            dry_run=True
        )

        assert results['success'] is False
        assert len(results['errors']) > 0
        assert "not found" in results['errors'][0].lower()

    def test_run_workflow_dry_run_basic(self):
        """Test workflow in dry-run mode."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create test workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["user1@example.com", "User1", "https://github.com/user/repo1", "Hello"])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is True
            assert results['total_entries'] == 1
            assert results['valid_entries'] == 1
            assert results['drafts_created'] == 1

    def test_run_workflow_multiple_entries(self):
        """Test workflow with multiple entries."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            for i in range(5):
                ws.append([
                    f"user{i}@example.com",
                    f"User{i}",
                    f"https://github.com/user/repo{i}",
                    f"Greeting{i}"
                ])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is True
            assert results['total_entries'] == 5
            assert results['valid_entries'] == 5
            assert results['drafts_created'] == 5

    def test_run_workflow_invalid_entries(self):
        """Test workflow with invalid entries."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["invalid-email", "User1", "https://github.com/user/repo1", "Hello"])
            ws.append(["user2@example.com", "User2", "https://github.com/user/repo2", "Hi"])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is True
            assert results['total_entries'] == 2
            assert results['valid_entries'] == 1
            assert results['invalid_entries'] == 1
            assert len(results['skipped_entries']) == 1

    def test_run_workflow_limit_entries(self):
        """Test workflow with entry limit."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            for i in range(10):
                ws.append([
                    f"user{i}@example.com",
                    f"User{i}",
                    f"https://github.com/user/repo{i}",
                    f"Greeting{i}"
                ])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True,
                limit=5
            )

            assert results['success'] is True
            assert results['total_entries'] == 10
            assert results['valid_entries'] == 10
            assert results['drafts_created'] == 5

    def test_run_workflow_missing_columns(self):
        """Test workflow with missing required columns."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name"])  # Missing Repo URL and greeting
            ws.append(["user@example.com", "User"])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is False
            assert len(results['errors']) > 0

    def test_run_workflow_result_structure(self):
        """Test that workflow returns correct result structure."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["user@example.com", "User", "https://github.com/user/repo", "Hello"])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            # Check structure
            assert "success" in results
            assert "total_entries" in results
            assert "valid_entries" in results
            assert "invalid_entries" in results
            assert "drafts_created" in results
            assert "drafts_failed" in results
            assert "errors" in results
            assert "skipped_entries" in results

    def test_run_workflow_mixed_valid_invalid(self):
        """Test workflow with mix of valid and invalid entries."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            # Valid
            ws.append(["user1@example.com", "User1", "https://github.com/user/repo1", "Hello"])
            # Invalid email
            ws.append(["invalid", "User2", "https://github.com/user/repo2", "Hi"])
            # Valid
            ws.append(["user3@example.com", "User3", "https://github.com/user/repo3", "Hey"])
            # Invalid URL
            ws.append(["user4@example.com", "User4", "not-a-url", "Greeting"])
            # Valid
            ws.append(["user5@example.com", "User5", "https://github.com/user/repo5", "Feedback"])
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is True
            assert results['total_entries'] == 5
            assert results['valid_entries'] == 3
            assert results['invalid_entries'] == 2
            assert results['drafts_created'] == 3

    def test_run_workflow_custom_delay(self):
        """Test workflow with custom API delay."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            ws.append(["user@example.com", "User", "https://github.com/user/repo", "Hello"])
            wb.save(tmp.name)

            # Should complete without error with custom delay
            results = run_workflow(
                input_file=tmp.name,
                dry_run=True,
                delay=0.1
            )

            assert results['success'] is True

    def test_run_workflow_empty_file(self):
        """Test workflow with empty Excel file (headers only)."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])
            # No data rows
            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            assert results['success'] is True
            assert results['total_entries'] == 0
            assert results['valid_entries'] == 0
            assert results['drafts_created'] == 0


class TestParseArguments:
    """Tests for parse_arguments function."""

    def test_parse_arguments_defaults(self):
        """Test argument parsing with defaults."""
        with patch('sys.argv', ['main.py']):
            args = parse_arguments()

            assert args.input == 'output/output34.xlsx'
            assert args.dry_run is False
            assert args.verbose is False
            assert args.limit is None
            assert args.delay == 0.5

    def test_parse_arguments_input_file(self):
        """Test argument parsing with input file."""
        with patch('sys.argv', ['main.py', '--input', 'custom.xlsx']):
            args = parse_arguments()

            assert args.input == 'custom.xlsx'

    def test_parse_arguments_dry_run(self):
        """Test argument parsing with dry-run flag."""
        with patch('sys.argv', ['main.py', '--dry-run']):
            args = parse_arguments()

            assert args.dry_run is True

    def test_parse_arguments_verbose(self):
        """Test argument parsing with verbose flag."""
        with patch('sys.argv', ['main.py', '--verbose']):
            args = parse_arguments()

            assert args.verbose is True

    def test_parse_arguments_limit(self):
        """Test argument parsing with limit."""
        with patch('sys.argv', ['main.py', '--limit', '10']):
            args = parse_arguments()

            assert args.limit == 10

    def test_parse_arguments_delay(self):
        """Test argument parsing with custom delay."""
        with patch('sys.argv', ['main.py', '--delay', '2.5']):
            args = parse_arguments()

            assert args.delay == 2.5

    def test_parse_arguments_all_options(self):
        """Test argument parsing with all options."""
        with patch('sys.argv', [
            'main.py',
            '--input', 'test.xlsx',
            '--dry-run',
            '--verbose',
            '--limit', '5',
            '--delay', '1.0'
        ]):
            args = parse_arguments()

            assert args.input == 'test.xlsx'
            assert args.dry_run is True
            assert args.verbose is True
            assert args.limit == 5
            assert args.delay == 1.0


class TestWorkflowIntegration:
    """Integration tests combining multiple components."""

    def test_end_to_end_workflow(self):
        """Test complete workflow from file reading to draft creation."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            # Create realistic input
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])

            test_data = [
                ("alice@example.com", "Alice Johnson", "https://github.com/alice/E18_ml_project", "Fantastic work!"),
                ("bob@example.com", "Bob Smith", "https://github.com/bob/E19_web_dev", "Great effort!"),
                ("charlie@example.com", "Charlie Brown", "https://github.com/charlie/E20_data_science", "Excellent submission!"),
            ]

            for email, name, repo_url, greeting in test_data:
                ws.append([email, name, repo_url, greeting])

            wb.save(tmp.name)

            # Run workflow
            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            # Verify
            assert results['success'] is True
            assert results['total_entries'] == 3
            assert results['valid_entries'] == 3
            assert results['drafts_created'] == 3
            assert results['drafts_failed'] == 0

    def test_workflow_with_partial_failures(self):
        """Test workflow handling partial failures gracefully."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["email", "name", "Repo URL", "greeting"])

            # Mix of valid and invalid
            ws.append(["valid@example.com", "Valid", "https://github.com/user/repo1", "Good"])
            ws.append(["invalid-email", "Invalid", "https://github.com/user/repo2", "Bad"])
            ws.append(["another@example.com", "Another", "https://github.com/user/repo3", "Nice"])

            wb.save(tmp.name)

            results = run_workflow(
                input_file=tmp.name,
                dry_run=True
            )

            # Should succeed overall, but report partial processing
            assert results['success'] is True
            assert results['total_entries'] == 3
            assert results['valid_entries'] == 2
            assert results['invalid_entries'] == 1
            assert results['drafts_created'] == 2
